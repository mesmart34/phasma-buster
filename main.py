import os
import re
from openpyxl import load_workbook

ghosts = []
evidences = []


def convert_string(s):
    # Convert the string to lower case and capitalize the first letter
    s = s.lower().capitalize()

    # Use regex to find patterns and replace them
    s = re.sub(r'_(\d+)_', lambda x: x.group(1), s)  # Remove underscores around digits
    s = re.sub(r'_([a-zA-Z])', lambda x: x.group(1).upper(), s)  # Remove underscore and capitalize the next letter

    return s


def generate_standard_evidences(sheet, row, header, class_code):
    name = convert_string(sheet.cell(row, 1).value)
    evidences.append(name)
    class_code += f'\tprivate static readonly StandardEvidence {name} = new ()\n'
    class_code += '\t{\n'
    class_code += f'\t\t{sheet.cell(header, 1).value} = PhasmaBusterTranslation.{sheet.cell(row, 1).value},\n'
    class_code += f'\t\t{sheet.cell(header, 2).value} = {sheet.cell(row, 2).value},\n'
    class_code += f'\t\t{sheet.cell(header, 3).value} = EvidenceType.{sheet.cell(row, 3).value},\n'
    class_code += f'\t\t{sheet.cell(header, 4).value} = Color.{sheet.cell(row, 4).value},\n'
    class_code += f'\t\t{sheet.cell(header, 5).value} = "{sheet.cell(row, 5).value}"\n'
    class_code += '\t};\n\n'
    return class_code


def tab(code: str, tab_count):
    result = ''
    for i in range(tab_count):
        result += '\t'
    return result


def add_sanity(class_code, sheet, row):
    class_code += '\t\tSanityEvidence = new ()\n'
    class_code += '\t\t{\n'

    class_code += f'\t\t\tName = PhasmaBusterTranslation.SANITY,\n'
    class_code += f'\t\t\tValues = new()\n'
    class_code += '\t\t\t{\n'
    sanity_evidences = str(sheet.cell(row, 5).value).split(', ')
    for sanity_evidence in sanity_evidences:
        class_code += '\t\t\t\tnew()\n'
        class_code += '\t\t\t\t{\n'
        class_code += f'\t\t\t\t\tSequence = {sanity_evidences.index(sanity_evidence)},\n'
        class_code += f'\t\t\t\t\tValue = {float(sanity_evidence)}m{(',\n' if sanity_evidences.index(sanity_evidence) < 2 else '\n')}'
        class_code += f'\t\t\t\t}}{(',\n' if sanity_evidences.index(sanity_evidence) < len(sanity_evidences) - 1 else '\n')}'
    class_code += '\t\t\t}\n'
    class_code += '\t\t},\n'
    return class_code


def add_speed(class_code, sheet, row):
    class_code += '\t\tSpeedEvidence = new ()\n'
    class_code += '\t\t{\n'

    class_code += f'\t\t\tName = PhasmaBusterTranslation.SPEED,\n'
    class_code += f'\t\t\tValues = new()\n'
    class_code += '\t\t\t{\n'
    sanity_evidences = str(sheet.cell(row, 6).value).split(', ')
    for sanity_evidence in sanity_evidences:
        class_code += '\t\t\t\tnew()\n'
        class_code += '\t\t\t\t{\n'
        class_code += f'\t\t\t\t\tSequence = {sanity_evidences.index(sanity_evidence)},\n'
        class_code += f'\t\t\t\t\tValue = {float(sanity_evidence)}m\n'
        class_code += f'\t\t\t\t}}{(',\n' if sanity_evidences.index(sanity_evidence) < len(sanity_evidences) - 1 else '\n')}'
    class_code += '\t\t\t}\n'
    class_code += '\t\t}\n'
    return class_code


def generate_hidden_evidences(sheet, class_code, row):
    global raw_owner
    class_code += '\t#region HiddenEvidences\n\n'
    row += 2
    while sheet.cell(row, 2).value is not None:
        if sheet.cell(row, 1).value is not None:
            raw_owner = sheet.cell(row, 1).value
            owner = convert_string(raw_owner)
        category = sheet.cell(row, 2).value + "_evidence"
        sequence = int(sheet.cell(row, 3).value)
        evidence_type = sheet.cell(row, 4).value
        file_path = sheet.cell(row, 5).value
        class_code += f'\tprivate static readonly {convert_string(category)} {f'{owner}Ev{sequence}'} = new ()\n'
        class_code += '\t{\n'
        class_code += f'\t\tName = PhasmaBusterTranslation.{raw_owner}_EV{sequence},\n'
        class_code += f'\t\tDescription = new ()\n'
        class_code += f'\t\t{{\n'
        class_code += f'\t\t\tText = {f'PhasmaBusterTranslation.{raw_owner}_EV{sequence}_D'},\n'
        if evidence_type is not None:
            class_code += f'\t\t\tEvidenceDescriptionType = EvidenceDescriptionType.{evidence_type},\n'
        if file_path is not None:
            class_code += f'\t\t\tFilePath = "{file_path}",\n'
        class_code += f'\t\t}},\n'
        if sheet.cell(row, 2).value == 'BEHAVIOUR':
            class_code += f'\t\tCategory = EvidenceType.Behaviours\n'
        else:
            class_code += f'\t\tCategory = EvidenceType.Tells\n'
        class_code += '\t};\n\n'
        row += 1
    class_code += '\t#endregion\n\n'
    return class_code


def add_standard_evidences(class_code, sheet, row):
    class_code += '\t\t{\n'
    class_code += '\t\t\t{\n'
    class_code += '\t\t\t\tEvidenceType.Standart,\n'
    class_code += '\t\t\t\tnew ()\n'
    class_code += '\t\t\t\t{\n'
    standard_evidences = str(sheet.cell(row, 2).value).split(', ')
    for standard_evidence in standard_evidences:
        class_code += f'\t\t\t\t\t{convert_string(standard_evidence)}{(',\n' if standard_evidences.index(standard_evidence) < 2 else '\n')}'
    class_code += '\t\t\t\t}\n'
    class_code += '\t\t\t},'
    return class_code


def add_tells_evidences(class_code, sheet, row, owner):
    tells_evidences = str(sheet.cell(row, 3).value).split(', ')
    if str(tells_evidences[0]) == 'None':
        return class_code
    class_code += f"""
            {{
                EvidenceType.Tells,
                new ()
                {{
"""
    
    for tells_evidence in tells_evidences:
        name = convert_string(f'{owner}_EV{tells_evidence}')
        evidences.append(name)
        class_code += f'\t\t\t\t\t{name}{(',\n' if tells_evidence.index(tells_evidence) < len(tells_evidence) else '\n')}'
    class_code += f"""
                }}
            }},"""
    return class_code


def add_behaviour_evidences(class_code, sheet, row, owner):
    behaviour_evidences = str(sheet.cell(row, 4).value).split(', ')
    if str(behaviour_evidences[0]) == 'None':
        return class_code
    class_code += f"""
            {{
                EvidenceType.Behaviours,
                new ()
                {{
"""

    for behaviour_evidence in behaviour_evidences:
        name = convert_string(f'{owner}_EV{behaviour_evidence}')
        evidences.append(name)
        class_code += f'\t\t\t\t\t{name}{(',\n' if behaviour_evidence.index(behaviour_evidence) < len(behaviour_evidence) else '\n')}'
    class_code += f"""
                }}
            }}\n"""
    return class_code


def add_evidences(class_code, sheet, row, owner):
    class_code += f'\t\tEvidences = new Dictionary<EvidenceType, List<BaseEvidence>>()\n'
    class_code = add_standard_evidences(class_code, sheet, row)
    class_code = add_tells_evidences(class_code, sheet, row, owner)
    class_code = add_behaviour_evidences(class_code, sheet, row, owner)
    class_code += '\t\t},\n'
    return class_code


def generate_ghost(sheet, row: int, class_code: str):
    name = convert_string(sheet.cell(row, 1).value)
    ghosts.append(name)
    class_code += f'\tprivate static readonly Ghost {name} = new ()\n'
    class_code += '\t{\n'
    class_code += f'\t\t{sheet.cell(1, 1).value} = PhasmaBusterTranslation.{sheet.cell(row, 1).value},\n'
    class_code = add_evidences(class_code, sheet, row, sheet.cell(row, 1).value)
    class_code = add_sanity(class_code, sheet, row)
    class_code = add_speed(class_code, sheet, row)

    class_code += '\t};\n\n'
    return class_code


def parse_standard_evidences(class_code, wb):
    sheet = wb['Evidences']
    header = 2
    row = 3
    class_code += '\t#region StandardEvidence\n\n'
    while row < 10:
        if sheet.cell(row, 1).value is None:
            break
        class_code = generate_standard_evidences(sheet, row, header, class_code)
        row += 1
    class_code += '\t#endregion\n\n'
    row += 1
    class_code = generate_hidden_evidences(sheet, class_code, row)
    return class_code


def parse_ghosts(class_code, wb):
    class_code += '\t#region Ghosts\n\n'
    sheet = wb['Ghosts']
    row = 2
    while True:
        if sheet.cell(row, 1).value is None:
            break
        class_code = generate_ghost(sheet, row, class_code)
        row += 1
    class_code += '\t#endregion\n\n'
    return class_code


def main():
    class_code = 'using System.Drawing;\n'
    class_code += 'using PhasmaBuster.Data.Common;\n'
    class_code += 'using PhasmaBuster.Data.Evidences;\n'
    class_code += 'using PhasmaBuster.Pages;\n\n'
    class_code += 'namespace PhasmaBuster.Data;\n\n'
    class_name = 'Model'
    class_code += f"public class {class_name}\n"
    class_code += '{\n'

    class_code += '\tpublic readonly Dictionary<BaseEvidence, EvidenceState> FilterEvidences = new();\n'
    class_code += '\tpublic readonly List<BaseEvidence> Evidences;\n'
    class_code += '\tpublic readonly List<Ghost> Ghosts;\n\n'

    wb = load_workbook('data.xlsx')

    class_code = parse_standard_evidences(class_code, wb)
    class_code = parse_ghosts(class_code, wb)

    class_code += '\tpublic Model()\n'
    class_code += '\t{\n'

    class_code += '\t\tEvidences = new()\n'
    class_code += '\t\t{\n'
    for evidence in evidences:
        class_code += f'\t\t\t{evidence},\n'
    class_code += '\t\t};\n'

    class_code += '\t\tGhosts = new()\n'
    class_code += '\t\t{\n'
    for ghost in ghosts:
        class_code += f'\t\t\t{ghost},\n'
    class_code += '\t\t};\n'

    class_code += '\t\tforeach (var baseEvidence in Evidences)\n'
    class_code += '\t\t{\n'
    class_code += '\t\t\tFilterEvidences.Add(baseEvidence, EvidenceState.Idle);\n'
    class_code += '\t\t}\n'

    class_code += '\t}\n'

    class_code += '}\n'
    output_file = f"{class_name}.cs"
    with open(output_file, 'w') as file:
        file.write(class_code)
    return


if __name__ == "__main__":
    main()
